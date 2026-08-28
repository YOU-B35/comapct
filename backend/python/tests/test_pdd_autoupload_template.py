# -*- coding: utf-8 -*-
"""拼多多批量上货模板契约测试。

CrossHub 自动上货模块的拼多多 Excel 会原样提交给 Commander 后端，
由 Commander 服务端 generic xlsx 解析器按固定列位解析（A~P 16 列：
internal/utils/xlsx.go -> XlsxData）。本测试锁定模板表头与解析器契约，
防止改表头导致真实后端解析错位。
"""

from pathlib import Path

import openpyxl

TEMPLATE = (
    Path(__file__).resolve().parents[3]
    / "dev"
    / "vue-site"
    / "public"
    / "templates"
    / "pinduoduo-publish-template.xlsx"
)

EXPECTED_HEADERS = [
    "产品图片",
    "产品类目",
    "状态",
    "编号",
    "型号",
    "数量",
    "变体价格(元)",
    "产品货号编码",
    "产品长（cm）",
    "产品宽（cm）",
    "产品高（cm）",
    "重量(g)",
    "备注",
    "产品标题中文（选填）",
    "产品标题英文（选填）",
    "产品标题俄文（选填）",
]


def test_template_exists():
    assert TEMPLATE.exists(), f"缺少拼多多模板: {TEMPLATE}"


def test_template_sheet_structure():
    wb = openpyxl.load_workbook(TEMPLATE)
    assert wb.sheetnames == ["选项", "填写说明", "上货数据"]


def test_data_sheet_headers_match_commander_parser():
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb["上货数据"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 17)]
    assert headers == EXPECTED_HEADERS


def test_help_sheet_mentions_pinduoduo_flow():
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb["填写说明"]
    text = "\n".join(
        str(ws.cell(row=r, column=1).value or "")
        for r in range(1, ws.max_row + 1)
    )
    assert "拼多多" in text
    assert "mms.pinduoduo.com" in text
    assert "存草稿" in text
